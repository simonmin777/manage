"""
not hard code, buy dynamicly read a work book
"""
import openpyxl
import datetime
import logging
import os
import shutil

ERROR_NOT_VALID_CYCLE = 701
ERROR_NOT_VALID_TENANT_CYCLE = 702
ERROR_NEGETIVE_NUMBER = 703
ERROR_NO_TENANT = 704

class Tenant:
    """ each tenant has attributs can be assigned """
    def __init__(self):
        """ default constructor, list all attributs and type """
        self.room = ' '     # str
        self.name = ' '     # str
        self.email = ' '    # str
        self.phone = ' '    # str
        self.movein = None          # datetime.datetime object
        self.moveout = None         # datetime.datetime object
        self.power_my_fee = 0.00    # float
        self.water_my_fee = 0.00    # float
        self.service_cycle = None   # ServiceCycle object
        self.service_power_days = 0       # int
        self.service_water_days = 0       # int
        self.sendemail = False      # bool

    def __repr__(self):
        rest = '%s | %s | %s ' % (self.room, self.name, self.email)
        rest += '\n' + str(self.service_cycle)
        rest += 'my power days = %d \tmy water days = %d\n' % (self.service_power_days, self.service_water_days)
        rest += 'my powew fee = %d \tmy water fee = %d\n\n' % (self.power_my_fee, self.water_my_fee)
        return rest

    def calculate_fees(self, power_all_days, water_all_days):
        if not power_all_days or not self.service_power_days:
            self.power_my_fee = 0
        else:
            self.power_my_fee = self.service_cycle.power_total_fee*self.service_power_days/power_all_days
            self.power_my_fee = int(self.power_my_fee*100)/100      # truncate after 2 decimal
        if not water_all_days or not self.service_water_days:
            self.water_my_fee = 0
        else:
            self.water_my_fee = self.service_cycle.water_total_fee*self.service_water_days/water_all_days
            self.water_my_fee = int(self.water_my_fee*100)/100      # truncate after 2 decimal

    def write_to_file(self, pathin='./housebill/'):
        path = './%s/' % pathin
        fout = open(path + "%s %s %s.txt" % (self.room, self.name, self.email) + '.txt', 'w')
        fout.write(self.get_email_txt())
        fout.close()

    def get_email_txt(self):
        # return a e-mail format string
        rest = 'Dear ' + self.name + ':\n\n'
        if self.service_cycle.is_power_cycle():
            rest += 'ENERGY Statement for the billing period from ' \
                    + str(self.service_cycle.power_start)[:10] + ' to ' \
                    + str(self.service_cycle.power_end)[:10] + '.\n\n'
            rest += 'You have been charged for %d days during this service period for a total of $%.02f.\n\n' \
                    % (self.service_power_days, self.power_my_fee)
        else:
            rest += 'There is no ENERGY Statement this billing period.\n\n'
        if self.service_cycle.is_water_cycle():
            rest += 'WATER Statement for the billing period from ' \
                    + str(self.service_cycle.water_start)[:10] + ' to ' \
                    + str(self.service_cycle.water_end)[:10] + '.\n\n'
            rest += 'You have been charged for %d days during this service period for a total of $%.02f.\n\n' \
                    % (self.service_water_days, self.water_my_fee)
        else:
            rest += 'WATER bill is sent every other month. no bill for this month.\n\n'
        # determine power + water
        rest += 'Total amount due is $%.02f + $%.02f = $%.02f.\n' % (self.power_my_fee, self.water_my_fee, self.power_my_fee + self.water_my_fee)
        rest += 'It is payable on the first day of the following month, i.e. %s.\n\n' % self.service_cycle.get_billday_string()[:10]
        rest += 'Make one single payment to cover the credit/debit forwarded from last month, utility bills and rent, if any.\n\n'
        rest += 'If you have any questions, please do not hesitate to ask.\n\n'
        rest += 'Best Regards,\nPaul\n\n'
        rest += 'A copy of the current statement is available upon request.\n'
        return rest


class ServiceCycle:
    """ contains mulitple (now only 2 service) cycle with fee"""
    def __init__(self, ws, rowindex, colindex):
        """ ROW COL extract ROW+1[COL, COL+1, COL+2] and ROW+2[COL, COL+1, COL+2] """
        self.power_start = ws[rowindex + 1][colindex].value
        self.power_end = ws[rowindex + 1][colindex + 1].value
        self.power_total_fee = ws[rowindex + 1][colindex + 2].value
        self.water_start = ws[rowindex + 2][colindex].value
        self.water_end = ws[rowindex + 2][colindex + 1].value
        self.water_total_fee = ws[rowindex + 2][colindex + 2].value

    def __repr__(self):
        rest = 'Service cyle:\n'
        if self.is_power_cycle():
            rest += 'Power: from %s to %s with fee %0.2f\n' % (str(self.power_start)[:10], str(self.power_end)[:10], self.power_total_fee)
        if self.is_water_cycle():
            rest += 'Water: from %s to %s with fee %0.2f\n' % (str(self.water_start)[:10], str(self.water_end)[:10], self.water_total_fee)
        return rest

    def is_valid_cycle(self):
        return self.is_power_cycle() or self.is_water_cycle()

    def is_power_cycle(self):
        return isinstance(self.power_start, datetime.datetime) and isinstance(self.power_end, datetime.datetime) and self.power_total_fee != 0

    def is_water_cycle(self):
        return isinstance(self.water_start, datetime.datetime) and isinstance(self.water_end, datetime.datetime) and self.water_total_fee != 0

    def get_power_days(self, movein, moveout):
        if moveout is None:
            moveout = self.power_end
        if not (isinstance(movein, datetime.datetime) and isinstance(moveout, datetime.datetime) and isinstance(self.power_start, datetime.datetime) and isinstance(self.power_end, datetime.datetime)):
            return 0
        elif self.power_end < movein or self.power_start > moveout:
            return 0
        else:
            return (min(moveout, self.power_end) - max(movein, self.power_start)).days + 1  # power days + 1

    def get_water_days(self, movein, moveout):
        delta = 0
        if moveout is None:
            moveout = self.water_end
        else:
            delta = 0   # water day by default not include last day, but moveout does

        if not (isinstance(movein, datetime.datetime) and isinstance(moveout, datetime.datetime) and isinstance(self.water_start, datetime.datetime) and isinstance(self.water_end, datetime.datetime)):
            return 0
        elif self.water_end < movein or self.water_start > moveout:
            return 0
        else:
            return (min(moveout, self.water_end) - max(movein, self.water_start)).days + delta

    def get_power_service_days(self):
        if self.is_power_cycle():
            return (self.power_end-self.power_start).days + 1  # power days + 1
        return 0

    def get_water_service_days(self):
        if self.is_water_cycle():
            return (self.water_end-self.water_start).days
        return 0

    def get_power_total_fee(self):
        if self.is_power_cycle():
            return self.power_total_fee
        return 0

    def get_water_total_fee(self):
        if self.is_water_cycle():
            return self.water_total_fee
        return 0

    def get_billday_string(self):
        """ return yyyy-mm-dd 10 char string """
        if not self.is_power_cycle():
            year = self.water_end.year
            month = self.water_end.month
        elif not self.is_water_cycle():
            year = self.power_end.year
            month = self.power_end.month
        else:
            year = max(self.power_end, self.water_end).year
            month = max(self.power_end, self.water_end).month
        month += 1
        if month == 13:
            year += 1
            month = 1
        return str(datetime.datetime(year, month, 1))[:10]


class Excel:
    """ to read an excel file with mulitple entry of tenant and service date """
    def __init__(self, filename, sheetname):
        """ construct key field lockation based on first row """
        self.wb = openpyxl.load_workbook(filename)  # must NOT to use read only mode
        self.ws = self.wb[sheetname]
        # read first row to determin key field
        matrix = {}
        _ = 0
        for _, cell in enumerate(self.ws[1]):
            if cell.value not in matrix and cell.value:
                matrix[str(cell.value).lower().lstrip().rstrip()] = cell.col_idx
            _ += 1
            if _ >= 26:
                break
        # extract key field, -1 because excel starts at 1
        self.room = matrix['room']-1
        self.name = matrix['tenants']-1
        self.email = matrix['e-mail']-1
        self.movein = matrix['move-in']-1
        self.moveout = matrix['move-out']-1
        self.service_dates = matrix['service dates']-1
        self.service_fee = matrix['fee']-1
        self.power_fee = matrix['power']-1
        self.water_fee = matrix['water']-1
        self.tenant = []
        # add send e-mail feature
        self.send_email = matrix['send e-mail']-1
        self.filename = filename        # for clean up, need save with filename

    @staticmethod
    def error_exit(string, error_code):
        logging.critical('ERROR %d - %s' % (error_code, string))
        exit(error_code)

    def tenant_check(self):
        billday = self.tenant[0].service_cycle.get_billday_string()
        for simon in self.tenant:
            # all billday must be the same
            if simon.service_cycle.get_billday_string() != billday:
                logging.error("Billday inconsistant %s vs %s" % (billday, simon.service_cycle.get_billday_string()))
            # days and fees must be 0 or positive
            if simon.service_power_days < 0 or simon.service_water_days < 0 or simon.power_my_fee < 0 or simon.water_my_fee < 0:
                Excel.error_exit(simon, ERROR_NEGETIVE_NUMBER)
            if simon.service_power_days == 0 and simon.service_water_days == 0:
                logging.info("[%s %s %s] is not an active tenant" % (simon.room, simon.name, simon.email))
            # warning if my days !=0 and < service cycle days
            else:
                if simon.service_power_days < simon.service_cycle.get_power_service_days():
                    logging.warning("[%s %s %s] only has %d power days out of %d" % (simon.room, simon.name, simon.email, simon.service_power_days, simon.service_cycle.get_power_service_days()))
                if simon.service_water_days < simon.service_cycle.get_water_service_days():
                    logging.warning("[%s %s %s] only has %d water days out of %d" % (simon.room, simon.name, simon.email, simon.service_water_days, simon.service_cycle.get_water_service_days()))

    def tenant_sum_check(self, tenant_index):
        """ print True of Fasle for sum of tenant [index : -1] """
        if len(self.tenant) == 0:
            Excel.error_exit('There is no tenant, why check sum?', ERROR_NO_TENANT)
        index = tenant_index
        tmp_power_sum = self.tenant[len(self.tenant)-1].service_cycle.get_power_total_fee()    # save last entry
        tmp_water_sum = self.tenant[len(self.tenant)-1].service_cycle.get_water_total_fee()
        tenant_power_sum = 0
        tenant_water_sum = 0
        while index < len(self.tenant):
            tenant_power_sum += self.tenant[index].power_my_fee
            tenant_water_sum += self.tenant[index].water_my_fee
            index += 1
        print('power %s: %.02f vs %.02f' % (str(int(100*abs(tenant_power_sum - tmp_power_sum)) <= len(self.tenant) - tenant_index), tenant_power_sum, tmp_power_sum))
        print('water %s: %.02f vs %.02f' % (str(int(abs(tenant_water_sum - tmp_water_sum)) <= len(self.tenant) - tenant_index), tenant_water_sum, tmp_water_sum))

    def is_valid_tenant_row(self, row):
        """ a valid tenant row must contains room, name, and email """
        if row[self.room].value is None or row[self.name].value is None or row[self.email].value is None:
                return False
        # and not 'service dates row'
        if row[self.service_dates].value is not None:
                return False
        return True

    def is_valid_service_dates_row(self, row):
        return str(row[self.service_dates].value).lower().lstrip().rstrip() == 'service dates'

    def load_tenant_from_row(self, row, service_cycle):
        """ return a Tenant object if sucessful, None otherwize """
        if not self.is_valid_tenant_row(row):
            return None
        simon = Tenant()
        simon.room = row[self.room].value
        simon.name = row[self.name].value
        simon.email = row[self.email].value
        simon.movein = row[self.movein].value
        simon.moveout = row[self.moveout].value
        simon.sendemail = str(row[self.send_email].value).lower().strip() == 'yes'
        simon.service_cycle = service_cycle
        simon.service_power_days = simon.service_cycle.get_power_days(simon.movein, simon.moveout)
        simon.service_water_days = simon.service_cycle.get_water_days(simon.movein, simon.moveout)
        if simon.service_power_days is None and simon.service_water_days is None:
            Excel.error_exit(list(map(lambda x: x.value, row)), ERROR_NOT_VALID_TENANT_CYCLE)
        # service fee only can be calculated after load all tenants (before new service cycle)
        return simon

    def divide_fees(self, power_all_days, water_all_days, tenant_start):
        """ all tenant from [tenant_start : -1] divide fees """
        index = tenant_start
        while index != len(self.tenant):
            self.tenant[index].calculate_fees(power_all_days, water_all_days)
            index += 1

    def process(self):
        """ process self.ws, all tenants should be load with correct fee """
        rowindex = 0
        service_cycle = None
        power_all_days = 0
        water_all_days = 0
        tenant_start = 0
        for row in self.ws:
            rowindex += 1
            # search a new service cycle
            if self.is_valid_service_dates_row(row):
                service_cycle = ServiceCycle(self.ws, rowindex, self.service_dates)
                if not service_cycle.is_valid_cycle():
                    Excel.error_exit('Row #%d is not a valid service cycle' % rowindex, ERROR_NOT_VALID_CYCLE)
                # after first row, every new service cycle need to calculate fees
                if rowindex > 1:
                    self.divide_fees(power_all_days, water_all_days, tenant_start)
                    self.tenant_sum_check(tenant_start)
                    tenant_start = len(self.tenant)
                    power_all_days = 0
                    water_all_days = 0
                # print(service_cycle)
                continue    # a new service cycle line cannot contain tenant
            # try to load row as tenant simon :P
            simon = self.load_tenant_from_row(row, service_cycle)
            if simon is not None:
                self.tenant.append(simon)
                power_all_days += simon.service_power_days
                water_all_days += simon.service_water_days
        # calculate the last service cycle fee
        self.divide_fees(power_all_days, water_all_days, tenant_start)
        self.tenant_sum_check(tenant_start)

    def backup(self, filename):
        """ backup eveything to a backfile and update fees """
        try:
            wbb = openpyxl.load_workbook(filename)      # must no use readonly
        except OSError:
            wbb = openpyxl.Workbook()                   # create new xlxs
            wbb.worksheets[0].title = 'Summary'
            wbb.worksheets[0].cell(1, 1).value = 'Bill date'
            wbb.worksheets[0].cell(1, 2).value = 'Utility'
            wbb.worksheets[0].cell(1, 3).value = 'Bank In'
            wbb.worksheets[0].cell(1, 4).value = 'Net In'
            wbb.worksheets[0].column_dimensions['a'].width = 15
            wbb.worksheets[0].column_dimensions['b'].width = 15
            wbb.worksheets[0].column_dimensions['c'].width = 15
            wbb.worksheets[0].column_dimensions['d'].width = 15
        wsb = wbb.create_sheet(self.tenant[0].service_cycle.get_billday_string())
        # add entry in summary sheet
        i = 2
        while wbb.worksheets[0].cell(i, 1).value is not None:
            i += 1
        wbb.worksheets[0].cell(i, 1).value = "%s" % wsb.title
        wbb.worksheets[0].cell(i, 2).value = "=SUM('%s'!K1:K200)" % wsb.title
        wbb.worksheets[0].cell(i, 3).value = "=SUM('%s'!R1:R200)" % wsb.title
        wbb.worksheets[0].cell(i, 4).value = "=C%d-B%d" % (i, i)
        # copy entire sheet first
        i, j = 1, 1
        for row in self.ws:
            for cell in row:
                wsb.cell(i, j).value = cell.value if not isinstance(cell.value, datetime.datetime) else str(cell.value)[0:10]
                j += 1
            i += 1
            j = 1
        # write fees for each tenant
        i = 1
        j = self.power_fee + 1
        k = self.water_fee + 1
        index = 0
        for row in self.ws:
            if self.is_valid_tenant_row(row):
                wsb.cell(i, j).value = self.tenant[index].power_my_fee
                wsb.cell(i, k).value = self.tenant[index].water_my_fee
                # add average daily value
                # if self.tenant[index].service_power_days:
                #     wsb.cell(i, j+4).value = self.tenant[index].power_my_fee / self.tenant[index].service_power_days
                # if self.tenant[index].service_water_days:
                #     wsb.cell(i, k+4).value = self.tenant[index].water_my_fee / self.tenant[index].service_water_days
                index += 1
            i += 1
        # add little bit style
        wsb.column_dimensions['b'].width = 20
        wsb.column_dimensions['c'].width = 25
        wsb.column_dimensions['d'].width = 15
        for _ in ['e', 'f', 'g', 'i', 'j']:
            wsb.column_dimensions[_].width = 12
        # for i in range(1, row):
        #     self.ws['d%d' % i].alignment = Alignment(horizontal='center')
        # save and close new workbookno
        wbb.save(filename)
        wbb.close()
        folder = '__backup__'
        if not os.path.isdir(folder):
            os.mkdir(folder)
            logging.info('Create directory %s' % folder)
        tenant_filename = filename[:-12]
        backup_filename = filename[:-5]
        surfix = '.xlsx'
        shutil.copy(tenant_filename+surfix, folder)
        shutil.copy(filename, folder)
        os.chdir(folder)
        # if filename+05 exist, remove it
        def remove_if_exist(path):
            if os.path.isfile(path):
                os.remove(path)
        remove_if_exist(tenant_filename+'05'+surfix)
        remove_if_exist(backup_filename+'05'+surfix)
        # rename each file 04->05, 03->04, until 00->01
        def rename_if_exist(name, num):
            if num == 0:
                path = name + surfix
                path2 = '%s01%s' % (name, surfix)
            else:
                path = '%s%02d%s' % (name, num, surfix)
                path2 = '%s%02d%s' % (name, num+1, surfix)
            if os.path.isfile(path):
                shutil.move(path, path2)
        for i in [4, 3, 2, 1, 0]:
            rename_if_exist(tenant_filename, i)
            rename_if_exist(backup_filename, i)
        os.chdir('..')

    def cleanup(self):
        """ clear every cell in colomn service_dates and fee to None if not service_dates """
        for row in self.ws:
            if not self.is_valid_service_dates_row(row):
                row[self.service_dates].value = None
                row[self.service_dates+1].value = None
                row[self.service_fee].value = None
        self.wb.save(self.filename)

    def write_all_tenant_to_file(self):
        folder = self.tenant[0].service_cycle.get_billday_string()[:10]
        if not os.path.isdir(folder):
            os.mkdir(folder)
            logging.info('Create directory %s' % folder)
        for simon in self.tenant:
            simon.write_to_file(folder)

    def clean_save(self):
        pass

    def close(self):
        """ manual close work sheet """
        self.wb.close()


# start top level
if __name__ == '__main__':
    def test():
        xlsx = Excel('SJ645.xlsx', 'next')
        xlsx.cleanup()
        xlsx.close()
    test()
# end of top level
